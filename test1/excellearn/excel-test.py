from openpyxl import load_workbook
import os
# 将表格中的数据打印出来

# 定位到excel的存放路径
current_path = os.path.abspath(__file__)
current_dir = os.path.split(current_path)[0]
exc_dir = os.path.join(current_dir, "excel_lab")
excel = os.path.join(exc_dir, "data.xlsx")

# 连接一个excel
wb = load_workbook(excel)
# 指向一个sheet
sh = wb["Sheet1"]
# 给单元格赋值1(赋值时，需要关闭被操作的excel，否则报错)
sh.cell(row=1, column=3).value = 2
# 给单元格赋值2
# sh["B2"] = 2

# 读出单元格内容
ce = sh.cell(row=2, column=5).value
print(ce)
max_column = sh.max_column
max_row = sh.max_row

wb.save(excel)
# sa = wb.active


# print(len_row)

# data =[]
# for i in range(1,max_row+1):
#     for j in range(1,max_column+1):
#         cl = sh.cell(row=i, column=j).value
#         data.append(cl)
#     print(data)


