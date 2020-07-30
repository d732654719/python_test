import logging, re
from openpyxl import load_workbook
from Common.path_config import *
class do_excel:
    def read_sheet(self,sheetname):
        wb = load_workbook(exc_path)
        sh = wb[sheetname]
        title_data = []
        for i in range(1, sh.max_column+1):
            title_data.append(sh.cell(row=1, column=i).value)
        test_data=[]
        for i in range(2, sh.max_row+1):
            dict = {}
            for j in range(1, sh.max_column+1):
                if sh.cell(row=i, column=j).value =="无":
                    dict[title_data[j-1]] = " "
                else:
                    if j == sh.max_column and sheetname == "测试用例":
                        dict[title_data[j-1]] = re.split("[,;\n]", sh.cell(row=i, column=j).value)
                    else:
                        dict[title_data[j-1]] = sh.cell(row=i, column=j).value
            test_data.append(dict)
        return test_data
    def read_config(self,sheetname):
        test_data = self.read_sheet("测试用例")
        test_cof = self.read_sheet("配置")[0]
        for case in test_data:
            case["接口地址"]=test_cof["服务器地址"]+case["接口地址"]
        #解决用例范围
        if test_cof["用例范围"]:
            sele_data = []
            try:
                if isinstance(eval(test_cof["用例范围"]),list):
                    for i in eval(test_cof["用例范围"]):
                        sele_data.append(test_data[i-1])
                elif isinstance(eval(test_cof["用例范围"]),tuple):
                    start = eval(test_cof["用例范围"])[0]
                    end = eval(test_cof["用例范围"][1])
                    for i in range(start,end+1):
                        sele_data.append(test_data[i - 1])
                else:
                    logging.error("[配置]-[用例范围]填写不规范，只可填列表和元组")
            except SyntaxError:
                logging.error("[配置]-[用例范围]语法填写错误，请重新填写")
            except IndexError:
                logging.error("[配置]-[用例范围]的用例编号超出范围")
        else:
            sele_data=test_data
        return sele_data


if __name__ == '__main__':
    data = do_excel().read_config("测试用例")
    print(data)

